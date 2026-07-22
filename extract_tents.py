#!/usr/bin/env python3
"""Extract the tent library from the Marknadsutstallare workbook into data/tents.json

    python3 scripts/extract_tents.py "Marknadsutstallare_2026-2.xlsx"

Reads the sheet "Utplaceringsdokument". Header is on ROW 2; data starts ROW 3.
Column map (1-indexed) as of the 2026-2 export:

     1 id (also carries the CATEGORY COLOUR as a cell fill)
     2 Kommentar                 (sensitive - never export)
     3 Nya                       1 = new vendor -> pink stripes
     4 Placering                 short letter code, e.g. MAT / MG / STH9
     5 Placering kommentar       (sensitive - never export)
     6 applicationName           display name
     7 contact_name              (sensitive - never export)
     8 tents                     JSON array of structures, may hold MORE THAN ONE
     9 Length                    metres (first structure only)
    10 Width                     metres (first structure only)
    11 electricity               '', 'none', '220v 6A', '380v 16A', '380v 32A'
    12 water                     'water' when the stand needs a supply
    13+ food_and_drink_vendor, special_requests, public_description,
        arrival_date/time, underlayments, ...   (sensitive - never export)

Two things that are easy to get wrong:

* Column A cell FILL is the only source of the category colour. openpyxl can read
  it; SheetJS in the browser cannot. Any browser-side re-import loses colours for
  ids it has never seen.
* ~24 rows have NO id - these are the organiser's own structures (Oltalt,
  Medeltidsveckans Taverna, Baldakin x2, Gute Rosteri x3, ...). They must be given
  stable synthetic ids or they collapse onto a single null id and the app treats
  them as one tent.
"""
import json
import pathlib
import re
import sys
import unicodedata
from collections import Counter

from openpyxl import load_workbook

SHEET = "Utplaceringsdokument"
FIRST_DATA_ROW = 3
COL = dict(id=1, nya=3, placering=4, name=6, tents=8, length=9, width=10,
           electricity=11, water=12)
DEFAULT_COLOUR = "#B6D7A8"


def fill_hex(cell):
    f = cell.fill
    if not f or f.patternType is None:
        return None
    c = f.fgColor
    if c and c.type == "rgb" and c.rgb and c.rgb != "00000000":
        return "#" + c.rgb[-6:]
    return None


def parse_arr(raw):
    """The `tents` cell is JSON: either a list or an object keyed "1","2",..."""
    if not raw:
        return []
    try:
        d = json.loads(raw)
    except (ValueError, TypeError):
        return []
    if isinstance(d, dict):
        d = list(d.values())
    return d if isinstance(d, list) else []


def dims(t):
    """Round structures are planned as a square of the diameter (user's call)."""
    if t.get("type") == "round":
        d = float(t.get("diameter") or 0)
        return (d, d) if d > 0 else (None, None)
    try:
        return (float(t.get("length") or 0) or None,
                float(t.get("width") or 0) or None)
    except (TypeError, ValueError):
        return (None, None)


def slug(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()[:28] or "item"


def extract(path):
    ws = load_workbook(path)[SHEET]
    seen, out, expanded, noid = Counter(), [], 0, 0

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        idv = ws.cell(row=r, column=COL["id"]).value
        name = ws.cell(row=r, column=COL["name"]).value
        if not name or not str(name).strip():
            continue
        nm = str(name).strip()

        if idv is None:                       # organiser structure
            noid += 1
            rid = "x_" + slug(nm)
            seen[rid] += 1
            if seen[rid] > 1:
                rid += f"_{seen[rid]}"
        else:
            rid = int(idv) if isinstance(idv, (int, float)) else idv

        def cell(key):
            v = ws.cell(row=r, column=COL[key]).value
            return str(v).strip() if v not in (None, "") else ""

        water = cell("water")
        base = {
            "id": rid,
            "name": nm,
            "placering": cell("placering"),
            "nya": 1 if ws.cell(row=r, column=COL["nya"]).value in (1, 1.0, "1") else 0,
            "color": fill_hex(ws.cell(row=r, column=COL["id"])) or DEFAULT_COLOUR,
            "electricity": cell("electricity"),
            "water": bool(water) and water.lower() not in ("no", "none", "0"),
        }

        L = ws.cell(row=r, column=COL["length"]).value
        W = ws.cell(row=r, column=COL["width"]).value
        arr = parse_arr(ws.cell(row=r, column=COL["tents"]).value)

        if len(arr) > 1:                      # vendor with several structures
            expanded += 1
            for i, t in enumerate(arr, 1):
                l, w = dims(t)
                rec = dict(base)
                rec["id"] = f"{base['id']}-{i}"
                rec["name"] = f"{nm} ({i}/{len(arr)})"
                rec["length"] = round(l, 2) if l else None
                rec["width"] = round(w, 2) if w else None
                out.append(rec)
        else:
            if (not L or not W) and arr:      # fall back to the JSON blob
                l, w = dims(arr[0])
                L, W = L or l, W or w
            rec = dict(base)
            rec["length"] = round(float(L), 2) if L else None
            rec["width"] = round(float(W), 2) if W else None
            out.append(rec)

    return out, expanded, noid


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    root = pathlib.Path(__file__).resolve().parent.parent
    out, expanded, noid = extract(sys.argv[1])

    dupes = [k for k, v in Counter(t["id"] for t in out).items() if v > 1]
    if dupes:
        print(f"ERROR: duplicate ids {dupes}", file=sys.stderr)
        return 1

    dest = root / "data" / "tents.json"
    dest.write_text(json.dumps(out, ensure_ascii=False), encoding="utf-8")
    print(f"{len(out)} entries -> {dest.relative_to(root)}")
    print(f"  multi-structure vendors expanded: {expanded}")
    print(f"  rows given synthetic ids: {noid}")
    print(f"  with size: {sum(1 for t in out if t['length'] and t['width'])}")
    print(f"  needing water: {sum(1 for t in out if t['water'])}")
    print(f"  colours: {dict(Counter(t['color'] for t in out))}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
